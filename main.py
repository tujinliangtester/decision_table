
from openpyxl import load_workbook

'''
设置单元格样式
'''


def judge_rule():
    fname = 'Decision_table.xlsx'
    wb = load_workbook(fname)

    sheet=wb['1']

    # excel行列从1开始
    for j in range(3,67):

        #使用了通用券，则不能使用单体券
        com_coupon=sheet.cell(row=7,column=j).value
        oli_coupon=sheet.cell(row=9,column=j).value
        no_oil_coupon=sheet.cell(row=11,column=j).value
        if (com_coupon == 'T'):
            if (oli_coupon=='T' or no_oil_coupon=='T' ):
                sheet.cell(row=5, column=j, value='need to kill')



    wb.save(fname)
    print('保存成功')


def activity_can_use():
    fname = 'Decision_table.xlsx'
    wb = load_workbook(fname)

    sheet = wb['1']
    for j in range(3, 67):
        com_plus = sheet.cell(row=36, column=j).value
        com_coupon=sheet.cell(row=37, column=j).value
        oil_plus=sheet.cell(row=38, column=j).value
        oil_coupon=sheet.cell(row=39, column=j).value
        no_oil_plus=sheet.cell(row=40, column=j).value
        no_oil_coupon=sheet.cell(row=41, column=j).value

        if(com_plus=='F' and com_coupon=='T'):
            sheet.cell(row=34, column=j, value='F')
            sheet.cell(row=35, column=j, value='F')
        if (oil_plus == 'F' and oil_coupon == 'T'):
            sheet.cell(row=34, column=j, value='F')
        if (no_oil_plus == 'F' and no_oil_coupon == 'T'):
            sheet.cell(row=35, column=j, value='F')

    wb.save(fname)
    print('保存成功')


def permutation_and_combination(index, conditional_pile_num, Condition_term_num=2):
    '''
    根据条件桩、条件项及条件的索引给出所以排列组合的情况
    :param index:条件的索引，注意，第一个索引为0
    :param conditional_pile_num:条件桩的数量
    :param Condition_term_num:每个条件项的取值可能性，默认是两个，即要么是F，要么是T
    :return:返回排列组合情况
    '''
    total_num= Condition_term_num ** conditional_pile_num
    s=''
    s_tmp=''
    s_tmp_length=total_num
    copy_times=index
    if(copy_times>0):
        s_tmp_length=total_num/(Condition_term_num**copy_times)
    for i in range(int(s_tmp_length/2)):
        s_tmp='T,'+s_tmp+'F,'
    s=s_tmp
    if(copy_times>0):
        for j in range(copy_times):
            s=s+s

    return s

def all_permutation_and_combination(conditional_pile_num, Condition_term_num=2):
    '''
    调用permutation_and_combination，返回所有的排列组合
    :param conditional_pile_num:条件桩的数量
    :param Condition_term_num:每个条件项的取值可能性，默认是两个，即要么是F，要么是T
    :return:所有的排列组合
    '''
    s_list=[]
    for i in range(conditional_pile_num):
        s_list.append(permutation_and_combination(index=i,conditional_pile_num=conditional_pile_num))
    for j in s_list:
        print(j)

    return s_list


if __name__ == '__main__':
    # s=permutation_and_combination(index=2,conditional_pile_num=3)
    # print('s:',s)
    # all_permutation_and_combination(conditional_pile_num=6)
    # judge_rule()
    activity_can_use()